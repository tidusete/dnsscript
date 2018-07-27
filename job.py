import csv


from openpyxl import load_workbook		#Lectura excel de dominios
wb = load_workbook(filename='LISTADO.xlsx')
#ws = wb['Poner nombre de la hoja']

colDominios = ws['A']

malDominis = list(colDominios)
dominis=[]
for x in range (5,len(malDominis)):
    dominis.append(malDominis[x].value.encode("utf-8"))


from subprocess import call
header = ["dominio"]
aux = []
for domini in dominis:
    try:
        call(["python", "dnsrecon/dnsrecon.py", "-d", domini, "-t", "std", "-c", "csv_"+domini])
        with open ("csv_"+domini,"rb") as csvfile:
            spamreader = csv.reader(csvfile, delimiter=',')
            tipus = []
            textAcumulat = []
            for row in spamreader:
                #print(row)
                trobat=False
                cont= 0
                while (not trobat and cont < len(tipus)):
                    if (row[0] not in header):
                        header.append(row[0])
                    if (row[0] == tipus[cont]):
                        trobat=True
                        textAcumulat[cont]=textAcumulat[cont]+" + "+row[1]+" "+row[2]
                    cont=cont+1
                if (not trobat):
                    tipus.append(row[0])
                    textAcumulat.append(row[1]+" "+row[2])
            aux2 = [domini,tipus,textAcumulat]
            aux.append(aux2)
        call(["rm", "csv_"+domini])
    except Exception:
        aux2 = [domini,[],[]]
        aux.append(aux2)

#abrimos csv final
trash=[]
with open('dnsfinal.csv', 'wb') as csvfile:
    spamwriter = csv.writer(csvfile, delimiter=",")
    spamwriter.writerow(header)
    for domini,nomdomini in enumerate(dominis):
        trash=[]

        trash.append(aux[domini][0])

        index=1
        while (index < len(header)):
            trobat3=False
            cont=1

            while (not trobat3 and cont < len(aux[domini][1])):
                if(aux[domini][1][cont] == header[index]):
                    trobat3=True
                    trash.append(aux[domini][2][cont])
                cont=cont+1
            if (not trobat3):
                trash.append(",")
            index=index+1
        spamwriter.writerow(trash)





